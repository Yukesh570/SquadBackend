from email import errors
import io
import os
from time import time
from urllib.parse import urljoin
from celery import shared_task
from django.apps import apps
from django.core.mail import send_mail
from django.core.mail import get_connection, EmailMessage
from django.core.mail import EmailMultiAlternatives
from datetime import datetime
import csv
import redis
from django.contrib.auth import get_user_model
from rest_framework import viewsets, status
from rest_framework.response import Response
import json
from django.core.files.storage import default_storage
from squadServices import models
from squadServices.helper.routeAndCostHelper import get_route_and_cost
from squadServices.models.campaign import CampaignContact
from squadServices.models.clientModel.client import Client
from squadServices.models.connectivityModel.verdor import Vendor
from squadServices.models.country import Country, Currency
from squadServices.models.finanace.invoice import ClientInvoice, VendorInvoice
from squadServices.models.finanace.invoiceSetup import InvoiceSetup
from squadServices.models.mappingSetup.mappingSetup import MappingSetup
from squadServices.models.operators.operators import OperatorNetworkCode, Operators
from squadServices.models.rateManagementModel.vendorRate import VendorRate
from squadServices.models.transaction.transaction import (
    ClientTransaction,
    TransactionType,
    VendorTransaction,
)
from squadServices.serializer.roleManagementSerializer.vendorRateSerializer import (
    VendorRateImportSerializer,
)
from django.db import transaction
from io import BytesIO
from django.template.loader import get_template
from django.core.files.base import ContentFile
from xhtml2pdf import pisa
from num2words import num2words
from dateutil import parser
import logging
import re  # Make sure this is at the top of your file!
import io

User = get_user_model()
from django.conf import settings
from squadServices.models.company import Company
from squadServices.models.email import EmailHost
import uuid
from django.db.models import Q

import html  # Add this to your imports at the top of the file!
import math
import openpyxl
from django.utils import timezone
from squadServices.models import Campaign
from squadServices.models.smpp.smppSMS import SMSMessage
from squadServices.models.detailedReport.detailedReport import DetailedSMSReport
from squadServices.helper.checkNumber import clean_phone_number
from squadServices.helper.smsSplitter import create_message_parts

logger = logging.getLogger(__name__)

redis_client = redis.StrictRedis.from_url(os.getenv("CELERY_RESULT_BACKEND"))


@shared_task
def sendEmailTask(
    subject, message, fromEmail, recipientList, emailHostId, attachments=None
):
    try:
        emailHost = EmailHost.objects.get(pk=emailHostId, isDeleted=False)
    except EmailHost.DoesNotExist:
        return
    if emailHost.security == "TLS":
        connection = get_connection(
            host=emailHost.smtpHost,
            port=emailHost.smtpPort,
            username=emailHost.smtpUser,
            password=emailHost.smtpPassword,
            use_tls=True,
        )
    elif emailHost.security == "SSL":
        connection = get_connection(
            host=emailHost.smtpHost,
            port=emailHost.smtpPort,
            username=emailHost.smtpUser,
            password=emailHost.smtpPassword,
            use_ssl=True,
        )
    email = EmailMultiAlternatives(
        subject=subject + "\u200b",
        body=message,
        from_email=fromEmail,
        to=recipientList,
        connection=connection,
    )

    email.content_subtype = "html"  # <-- tell Django this is HTML
    if attachments:
        for att in attachments:
            """
            att must contain:
            {
                "name": "invoice.pdf",
                "content": "<BASE64_STRING>",
                "type": "application/pdf"
            }
            """
            import base64

            file_name = att["name"]
            file_content = base64.b64decode(att["content"])
            mime_type = att.get("type", "application/octet-stream")

            email.attach(file_name, file_content, mime_type)

    email.send()


@shared_task()
def export_model_csv(model_name: str, filters=None, fields=None, module=None):
    """
    Generate CSV for any model asynchronously.
    """
    print(model_name, filters, fields, module)

    task_id = export_model_csv.request.id  # unique task ID

    # Resolve model
    try:
        Model = apps.get_model(model_name)
    except LookupError:
        redis_client.set(task_id, "error:Model not found")

        return f"Model {model_name} not found"

    # Generate unique filename
    filename = (
        f"{module or Model._meta.model_name}_{uuid.uuid4().hex}_{int(time())}.csv"
    )
    export_dir = os.path.join(settings.MEDIA_ROOT, "exports")
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)

    # Base queryset
    queryset = (
        Model.objects.filter(isDeleted=False)
        if hasattr(Model, "isDeleted")
        else Model.objects.all()
    )

    # Apply filters
    if filters:
        transformed_filters = {f"{k}__icontains": v for k, v in filters.items()}
        queryset = queryset.filter(**transformed_filters)
    total = queryset.count()
    if total == 0:
        redis_client.set(task_id, "100")
        redis_client.set(f"{task_id}_result", filename)
        return filename
    # Determine fields
    if fields is None:
        fields = [field.name for field in Model._meta.fields]
    processed = 0

    # Write CSV
    # Write CSV
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(fields)

        for obj in queryset.iterator(chunk_size=2000):
            processed += 1

            row = []
            for field_name in fields:
                val = getattr(obj, field_name, None)

                # Safely extract the 'name' if it's a Foreign Key
                if val is not None and hasattr(val, "name"):
                    val = val.name
                elif val is None:
                    val = ""
                else:
                    val = str(val)

                # THE EXCEL HACK:
                # If the string is only digits and commas (like "734,901"),
                # wrap it in an Excel formula so Excel doesn't delete the comma.
                if re.fullmatch(r"[\d,]+", val) and "," in val:
                    val = f'="{val}"'

                row.append(val)

            writer.writerow(row)

            # Update Redis
            progress = int((processed / total) * 100)
            redis_client.set(task_id, str(progress))
    # Schedule deletion after 5 minutes
    deleteForLater.apply_async(args=[filepath], countdown=300)
    redis_client.set(task_id, "100")
    redis_client.set(f"{task_id}_result", filename)
    return filename


@shared_task
def deleteForLater(filepath):
    """Delete exported CSV file safely."""
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
    except Exception as e:
        print(f"Failed to delete {filepath}: {e}")


@shared_task(bind=True)
def import_vendor_rate_task(self, filepath, user_id, task_id, mapping_id):
    redis_client.set(task_id, "0")

    created = 0
    failed = []

    mapping = MappingSetup.objects.filter(id=mapping_id).first()
    if not mapping:
        redis_client.set(task_id, "error: Mapping not found")
        if os.path.exists(filepath):
            os.remove(filepath)
        return

    # Normalize header_map once
    header_map = {
        mapping.ratePlan.lower().strip(): "ratePlan",
        mapping.country.lower().strip(): "country",
        mapping.countryCode.lower().strip(): "countryCode",
        mapping.timeZone.lower().strip(): "timeZone",
        mapping.network.lower().strip(): "network",
        mapping.MCC.lower().strip(): "MCC",
        mapping.MNC.lower().strip(): "MNC",
        mapping.rate.lower().strip(): "rate",
        mapping.dateTime.lower().strip(): "dateTime",
    }

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)

            # Pre-map CSV header names -> model fields
            csv_header_map = {}
            for col in reader.fieldnames:
                normalized_col = col.lower().strip()
                logger.info(f"normalized_col={normalized_col}")
                logger.info(f"header_map={header_map}")

                if normalized_col in header_map:
                    csv_header_map[col] = header_map[normalized_col]

            rows = list(reader)
            total = len(rows)
            user = User.objects.get(id=user_id)
            print("=========!!!!====================")
            # ⚡️ 1. OUTER TRANSACTION: Opens the temporary workspace
            with transaction.atomic():
                for index, row in enumerate(rows, start=1):

                    mapped_row = {}

                    # Use pre-mapped header mapping
                    for col, val in row.items():
                        if col in csv_header_map:
                            mapped_row[csv_header_map[col]] = val

                    # Clean numeric fields
                    for key in ["MCC", "MNC", "rate", "countryCode"]:
                        if mapped_row.get(key) == "":
                            mapped_row[key] = None

                    # Parse datetime
                    if mapped_row.get("dateTime"):
                        try:
                            mapped_row["dateTime"] = parser.parse(
                                mapped_row["dateTime"]
                            )
                        except:
                            failed.append({"row": index, "error": "Invalid dateTime"})
                            continue
                    # Example before serializer.save()
                    if VendorRate.objects.filter(
                        ratePlan=mapped_row.get("ratePlan"),
                    ).exists():
                        failed.append({"row": index, "error": "Duplicate entry"})
                        continue

                    serializer = VendorRateImportSerializer(data=mapped_row)

                    if serializer.is_valid():
                        try:
                            # ⚡️ 2. INNER TRANSACTION: Opens the actual transaction
                            with transaction.atomic():
                                serializer.save(createdBy=user, updatedBy=user)
                                created += 1
                        except Exception as e:
                            failed.append({"row": index, "error": str(e)})
                    else:
                        failed.append({"row": index, "error": serializer.errors})

                    # Update progress
                    progress = int((index / total) * 100)
                    redis_client.set(task_id, str(progress))
                if failed:
                    transaction.set_rollback(True)
                    final_message = (
                        f"Import failed. 0 of {total} rows saved due to errors."
                    )
                else:
                    final_message = f"Success! {total} rows processed and saved."

        redis_client.set(task_id, "100")
        redis_client.set(
            f"{task_id}_result",
            json.dumps({"message": final_message, "errors": failed}),
        )
        deleteForLater.apply_async(args=[filepath], countdown=100)
        if os.path.exists(filepath):
            os.remove(filepath)

        return "done"
    except Exception as e:
        redis_client.set(task_id, f"error:{str(e)}")
        deleteForLater.apply_async(args=[filepath], countdown=100)
        return


@shared_task(bind=True)
def import_country_task(self, filepath, user_id, task_id):
    """
    Import countries from CSV.
    Expected CSV headers: name,countryCode,MCC
    """
    redis_client.set(task_id, "0")
    errors = []

    try:
        user = User.objects.filter(id=user_id).first()

        with open(filepath, newline="", encoding="utf-8") as csvfile:
            rows = list(csv.DictReader(csvfile))
            total = len(rows)

            if total == 0:
                redis_client.set(task_id, "100")
                redis_client.set(
                    f"{task_id}_result",
                    json.dumps({"message": "Empty CSV", "errors": []}),
                )
                return

            # ⚡️ 1. OUTER TRANSACTION: Opens the temporary workspace
            with transaction.atomic():
                for index, row in enumerate(rows, start=1):
                    name = row.get("name", "").strip()
                    code = row.get("countryCode", "").strip()

                    if not name or not code:
                        errors.append(
                            {"row": index, "error": "Missing name or countryCode"}
                        )
                        continue

                    # Check if country already exists (by name)
                    if Country.objects.filter(
                        Q(name__iexact=name), isDeleted=False
                    ).exists():
                        errors.append(
                            {
                                "row": index,
                                "error": f"Country '{name}' already exists",
                            }
                        )
                        continue

                    raw_mcc = str(row.get("MCC") or "0").strip()

                    try:
                        # ⚡️ 2. INNER TRANSACTION: Catches DB-level crashes without breaking the loop
                        with transaction.atomic():
                            Country.objects.create(
                                name=name,
                                countryCode=code,
                                MCC=raw_mcc,
                                createdBy=user,
                                updatedBy=user,
                                isDeleted=False,
                            )
                    except Exception as e:
                        errors.append({"row": index, "error": str(e)})
                        continue

                    # Update progress
                    progress = int((index / total) * 100)
                    redis_client.set(task_id, str(progress))

                # ⚡️ 3. ROLLBACK TRIGGER: Deletes everything if even one error was found
                if errors:
                    transaction.set_rollback(True)
                    final_message = (
                        f"Import failed. 0 of {total} rows saved due to errors."
                    )
                else:
                    final_message = f"Success! {total} rows processed and saved."

        # Mark complete
        redis_client.set(task_id, "100")
        redis_client.set(
            f"{task_id}_result",
            json.dumps({"message": final_message, "errors": errors}),
        )

        deleteForLater.apply_async(args=[filepath], countdown=100)

    except Exception as e:
        redis_client.set(task_id, f"error:{str(e)}")
        deleteForLater.apply_async(args=[filepath], countdown=100)


@shared_task(bind=True)
def import_operator_task(self, filepath, user_id, task_id):
    """
    Import operators from CSV.
    Expected CSV headers: name,country,MNC
    """
    redis_client.set(task_id, "0")  # initial progress
    errors = []

    try:
        with open(filepath, newline="", encoding="utf-8") as csvfile:
            rows = list(csv.DictReader(csvfile))
            total = len(rows)
            user = User.objects.get(id=user_id)

            if total == 0:
                redis_client.set(task_id, "100")
                redis_client.set(
                    f"{task_id}_result",
                    json.dumps({"message": "Empty CSV", "errors": []}),
                )
                return

            for index, row in enumerate(rows, start=1):
                name = row.get("name", "").strip()
                country_name = row.get("country", "").strip()
                mnc = row.get("MNC")

                # Validate required fields
                if not name or not country_name:
                    errors.append({"row": index, "error": "Missing name or country"})
                    continue

                # Check if country exists
                country = Country.objects.filter(
                    name__iexact=country_name, isDeleted=False
                ).first()
                if not country:
                    errors.append(
                        {"row": index, "error": f"Country '{country_name}' not found"}
                    )
                    continue

                # Check if operator already exists
                if Operators.objects.filter(
                    name=name, country=country, isDeleted=False
                ).exists():
                    errors.append(
                        {
                            "row": index,
                            "error": f"Operator '{name}' for country '{country_name}' already exists",
                        }
                    )
                    continue

                try:
                    Operators.objects.create(
                        name=name,
                        country=country,
                        MNC=int(mnc) if mnc else 0,
                        createdBy=user,
                        updatedBy=user,
                        isDeleted=False,
                    )
                except Exception as e:
                    errors.append({"row": index, "error": str(e)})
                    continue

                # Update progress
                progress = int((index / total) * 100)
                redis_client.set(task_id, str(progress))

        # Mark complete
        redis_client.set(task_id, "100")
        redis_client.set(
            f"{task_id}_result",
            json.dumps({"message": f"{total} rows processed", "errors": errors}),
        )

        # Schedule file deletion
        deleteForLater.apply_async(args=[filepath], countdown=100)

    except Exception as e:
        redis_client.set(task_id, f"error:{str(e)}")
        deleteForLater.apply_async(args=[filepath], countdown=100)


@shared_task(bind=True)
def import_currency_task(self, filepath, user_id, task_id):
    """
    Import currencies from CSV.
    Expected CSV headers: name,currencyCode,numericCode,symbol,decimalPlaces,isActive
    """
    redis_client.set(task_id, "0")  # initial progress
    errors = []

    try:
        # ⚡️ 1. Read the file as raw bytes first
        with open(filepath, "rb") as f:
            raw_data = f.read()

        # ⚡️ 2. Smart Decoding
        try:
            # Try UTF-8 first (using -sig to strip Excel's hidden BOM characters)
            decoded_content = raw_data.decode("utf-8-sig")
        except UnicodeDecodeError:
            # Fallback to standard Excel Windows encoding if a € or £ trips up UTF-8
            decoded_content = raw_data.decode("windows-1252")

        # ⚡️ 3. Pass the successfully decoded string into the CSV DictReader
        csvfile = io.StringIO(decoded_content)
        rows = list(csv.DictReader(csvfile))
        total = len(rows)

        user = User.objects.get(id=user_id)

        if total == 0:
            redis_client.set(task_id, "100")
            redis_client.set(
                f"{task_id}_result",
                json.dumps({"message": "Empty CSV", "errors": []}),
            )
            return

        with transaction.atomic():
            for index, row in enumerate(rows, start=1):
                name = row.get("name", "").strip()
                currency_code = row.get("currencyCode", "").strip()
                numeric_code = row.get("numericCode", "").strip()
                symbol = row.get("symbol", "").strip()

                # Handle decimal places (default to 2 if empty or invalid)
                raw_decimal = row.get("decimalPlaces", "").strip()
                try:
                    decimal_places = int(raw_decimal) if raw_decimal else 2
                except ValueError:
                    decimal_places = 2

                # Handle boolean conversion for isActive (defaults to True)
                raw_is_active = str(row.get("isActive", "true")).strip().lower()
                is_active = raw_is_active in ["true", "1", "yes", "y"]

                # Validate required fields
                if not name:
                    errors.append({"row": index, "error": "Missing currency name"})
                    continue

                # Check if currency already exists (checking by name)
                if Currency.objects.filter(name__iexact=name, isDeleted=False).exists():
                    errors.append(
                        {
                            "row": index,
                            "error": f"Currency '{name}' already exists",
                        }
                    )
                    continue

                try:
                    with transaction.atomic():
                        Currency.objects.create(
                            name=name,
                            currencyCode=currency_code if currency_code else None,
                            numericCode=numeric_code if numeric_code else None,
                            symbol=symbol if symbol else None,
                            decimalPlaces=decimal_places,
                            isActive=is_active,
                            isDeleted=False,
                            createdBy=user,
                            updatedBy=user,
                        )
                except Exception as e:
                    errors.append({"row": index, "error": str(e)})
                    continue

                # Update progress
                progress = int((index / total) * 100)
                redis_client.set(task_id, str(progress))

            # THE ROLLBACK TRIGGER
            if errors:
                transaction.set_rollback(True)
                final_message = f"Import failed. 0 of {total} rows saved due to errors."
            else:
                final_message = f"Success! {total} rows processed and saved."

        # Mark complete
        redis_client.set(task_id, "100")
        redis_client.set(
            f"{task_id}_result",
            json.dumps({"message": final_message, "errors": errors}),
        )

        # Schedule file deletion
        deleteForLater.apply_async(args=[filepath], countdown=100)

    except Exception as e:
        redis_client.set(task_id, f"error:{str(e)}")
        deleteForLater.apply_async(args=[filepath], countdown=100)


@shared_task(bind=True)
def import_operator_network_code_task(self, filepath, user_id, task_id):
    """
    Import Operator Network Codes from CSV.
    Expected CSV headers: operator_name, country_name, MCC, MNC, networkName, networkType, isPrimary, status, effectiveFrom, effectiveTo, notes
    Dates must be in YYYY-MM-DD format.
    """
    redis_client.set(task_id, "0")  # initial progress
    errors = []

    VALID_NETWORK_TYPES = ["GSM", "LTE", "5G", "CDMA", "UNKNOWN"]
    VALID_STATUSES = ["ACTIVE", "INACTIVE"]

    try:
        with open(filepath, newline="", encoding="utf-8") as csvfile:
            rows = list(csv.DictReader(csvfile))
            total = len(rows)
            user = User.objects.get(id=user_id)

            if total == 0:
                redis_client.set(task_id, "100")
                redis_client.set(
                    f"{task_id}_result",
                    json.dumps({"message": "Empty CSV", "errors": []}),
                )
                return

            # ⚡️ OPEN THE MASTER TRANSACTION BLOCK
            with transaction.atomic():
                for index, row in enumerate(rows, start=1):
                    # 1. Extract strings
                    operator_name = row.get("operator", "").strip()
                    country_name = row.get("country", "").strip()
                    mcc = row.get("MCC", "").strip()
                    mnc = row.get("MNC", "").strip()
                    network_name = row.get("networkName", "").strip()
                    notes = row.get("notes", "").strip()

                    # 2. Extract and sanitize Choice Fields
                    raw_network_type = row.get("networkType", "").strip().upper()
                    network_type = (
                        raw_network_type
                        if raw_network_type in VALID_NETWORK_TYPES
                        else "UNKNOWN"
                    )

                    raw_status = row.get("status", "").strip().upper()
                    status = raw_status if raw_status in VALID_STATUSES else "ACTIVE"

                    # 3. Extract and parse Boolean
                    raw_is_primary = str(row.get("isPrimary", "false")).strip().lower()
                    is_primary = raw_is_primary in ["true", "1", "yes", "y"]

                    # 4. Extract and parse Dates (YYYY-MM-DD)
                    effective_from = None
                    effective_to = None

                    try:
                        raw_from = row.get("effectiveFrom", "").strip()
                        if raw_from:
                            effective_from = datetime.strptime(
                                raw_from, "%Y-%m-%d"
                            ).date()

                        raw_to = row.get("effectiveTo", "").strip()
                        if raw_to:
                            effective_to = datetime.strptime(raw_to, "%Y-%m-%d").date()
                    except ValueError:
                        errors.append(
                            {
                                "row": index,
                                "error": "Invalid date format. Must be YYYY-MM-DD.",
                            }
                        )
                        continue

                    # 5. Validate mandatory fields
                    if not operator_name or not country_name or not mcc or not mnc:
                        errors.append(
                            {
                                "row": index,
                                "error": "Missing operator_name, country_name, MCC, or MNC",
                            }
                        )
                        continue

                    # 6. Look up Foreign Keys
                    country = Country.objects.filter(
                        name__iexact=country_name, isDeleted=False
                    ).first()
                    if not country:
                        errors.append(
                            {
                                "row": index,
                                "error": f"Country '{country_name}' not found",
                            }
                        )
                        continue

                    operator_obj = Operators.objects.filter(
                        name__iexact=operator_name, isDeleted=False
                    ).first()
                    if not operator_obj:
                        errors.append(
                            {
                                "row": index,
                                "error": f"Operator '{operator_name}' not found",
                            }
                        )
                        continue

                    # 7. Check Unique Together constraint
                    if OperatorNetworkCode.objects.filter(
                        operator=operator_obj, MCC=mcc, MNC=mnc, isDeleted=False
                    ).exists():
                        errors.append(
                            {
                                "row": index,
                                "error": f"Network Code MCC{mcc}-MNC{mnc} for Operator '{operator_name}' already exists",
                            }
                        )
                        continue

                    # 8. Create the record
                    try:
                        # ⚡️ INNER ATOMIC BLOCK: Prevents Django from crashing the outer transaction if this specific create fails
                        with transaction.atomic():
                            OperatorNetworkCode.objects.create(
                                operator=operator_obj,
                                country=country,
                                MCC=mcc,
                                MNC=mnc,
                                networkName=network_name if network_name else None,
                                networkType=network_type,
                                isPrimary=is_primary,
                                status=status,
                                effectiveFrom=effective_from,
                                effectiveTo=effective_to,
                                notes=notes if notes else None,
                                createdBy=user,
                                updatedBy=user,
                                isDeleted=False,
                            )
                    except Exception as e:
                        errors.append({"row": index, "error": str(e)})
                        continue

                    # Update progress
                    progress = int((index / total) * 100)
                    redis_client.set(task_id, str(progress))

                # ⚡️ THE ALL-OR-NOTHING LOGIC
                # Once all rows are processed, we check if ANY errors were collected
                if errors:
                    transaction.set_rollback(
                        True
                    )  # This deletes all the objects we just created!
                    final_message = (
                        f"Import failed. 0 of {total} rows saved due to errors."
                    )
                else:
                    final_message = f"Success! {total} rows processed and saved."

        # Mark complete and save the final report to Redis
        redis_client.set(task_id, "100")
        redis_client.set(
            f"{task_id}_result",
            json.dumps({"message": final_message, "errors": errors}),
        )

        # Schedule file deletion
        deleteForLater.apply_async(args=[filepath], countdown=100)

    except Exception as e:
        redis_client.set(task_id, f"error:{str(e)}")
        deleteForLater.apply_async(args=[filepath], countdown=100)


@shared_task
def generate_invoice_pdf_task(invoice_id, breakdown_data, tax_amount=0, base_url=""):
    """
    Background task to render HTML, convert to PDF, and save it to the database.
    """
    print(f"Generating PDF for invoice {invoice_id}")
    try:
        # 1. Fetch the invoice from the DB using the ID passed to Celery
        invoice_obj = ClientInvoice.objects.get(id=invoice_id)
    except ClientInvoice.DoesNotExist:
        print(f"Error: Invoice {invoice_id} not found.")
        return False
    client_company = invoice_obj.client.company
    setup_rules = InvoiceSetup.objects.filter(
        company=client_company, isDeleted=False
    ).first()
    business_entity = setup_rules.businessEntity if setup_rules else None

    if setup_rules and setup_rules.billingAddressOverride:
        final_address = setup_rules.billingAddressOverride
    else:
        final_address = client_company.address
    # 2. Calculate the final math
    total_amount = invoice_obj.totalAmount
    grand_total = float(total_amount) + float(tax_amount)

    # 3. Convert numbers to words
    amount_in_words = num2words(int(grand_total))
    # Prepare the logo URL safely
    logo_url = None
    if business_entity and business_entity.companyLogo:
        if base_url:
            # This safely combines "http://domain.com/" and "/media/logos/img.png"
            logo_url = urljoin(base_url, business_entity.companyLogo.url)
        else:
            logo_url = business_entity.companyLogo.url
    # 4. Prepare Context
    context = {
        "client": invoice_obj.client,
        "entity_name": (business_entity.companyName if business_entity else "N/A"),
        "entity_logo": logo_url,
        "entity_address": (
            business_entity.businessAddress if business_entity else "N/A"
        ),
        "entity_email": (business_entity.emailAddress if business_entity else "N/A"),
        "entity_phone": (business_entity.phone if business_entity else "N/A"),
        "client_name": client_company.name,
        "client_email": client_company.companyEmail,
        "client_currency": client_company.currency,
        "client_phone": client_company.phone,
        "client_address": final_address,
        "breakdown": breakdown_data,
        "total_amount": total_amount,
        "tax_amount": tax_amount,
        "grand_total": grand_total,
        "amount_in_words": amount_in_words,
        "bank_details": business_entity.bankAccountDetail,
    }

    # 5. Render HTML
    template = get_template("finance/invoice_pdf.html")
    html_string = template.render(context)

    # 6. Convert to PDF
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(BytesIO(html_string.encode("UTF-8")), dest=pdf_buffer)

    # 7. Save to the Database
    if not pisa_status.err:
        pdf_name = f"{invoice_obj.invoiceNumber}.pdf"
        invoice_obj.invoicePdf.save(pdf_name, ContentFile(pdf_buffer.getvalue()))
        invoice_obj.save()
        return f"Successfully generated PDF for Invoice {invoice_obj.invoiceNumber}"

    return f"Failed to generate PDF for Invoice {invoice_obj.invoiceNumber}"


@shared_task
def generate_vendorInvoice_pdf_task(
    invoice_id, breakdown_data, tax_amount=0, base_url=""
):
    """
    Background task to render HTML, convert to PDF, and save it to the database.
    """
    print(f"Generating PDF for invoice {invoice_id}")
    try:
        # 1. Fetch the invoice from the DB using the ID passed to Celery
        invoice_obj = VendorInvoice.objects.get(id=invoice_id)
    except VendorInvoice.DoesNotExist:
        print(f"Error: Invoice {invoice_id} not found.")
        return False
    vendor_company = invoice_obj.vendor.company
    setup_rules = InvoiceSetup.objects.filter(
        company=vendor_company, isDeleted=False
    ).first()
    business_entity = setup_rules.businessEntity if setup_rules else None

    if setup_rules and setup_rules.billingAddressOverride:
        final_address = setup_rules.billingAddressOverride
    else:
        final_address = vendor_company.address
    # 2. Calculate the final math
    total_amount = invoice_obj.totalAmount
    grand_total = float(total_amount) + float(tax_amount)

    # 3. Convert numbers to words
    amount_in_words = num2words(int(grand_total))
    logo_url = None
    if business_entity and business_entity.companyLogo:
        if base_url:
            # This safely combines "http://domain.com/" and "/media/logos/img.png"
            logo_url = urljoin(base_url, business_entity.companyLogo.url)
        else:
            logo_url = business_entity.companyLogo.url
    # 4. Prepare Context
    context = {
        "vendor": invoice_obj.vendor,
        "entity_name": (business_entity.companyName if business_entity else "N/A"),
        "entity_logo": logo_url,
        "entity_address": (
            business_entity.businessAddress if business_entity else "N/A"
        ),
        "entity_email": (business_entity.emailAddress if business_entity else "N/A"),
        "entity_phone": (business_entity.phone if business_entity else "N/A"),
        "vendor_name": vendor_company.name,
        "vendor_email": vendor_company.companyEmail,
        "vendor_currency": vendor_company.currency,
        "vendor_phone": vendor_company.phone,
        "vendor_address": final_address,
        "breakdown": breakdown_data,
        "total_amount": total_amount,
        "tax_amount": tax_amount,
        "grand_total": grand_total,
        "amount_in_words": amount_in_words,
        "bank_details": business_entity.bankAccountDetail if business_entity else "N/A",
    }

    # 5. Render HTML
    template = get_template("finance/vendorInvoice_pdf.html")
    html_string = template.render(context)

    # 6. Convert to PDF
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(BytesIO(html_string.encode("UTF-8")), dest=pdf_buffer)

    # 7. Save to the Database
    if not pisa_status.err:
        pdf_name = f"{invoice_obj.invoiceNumber}.pdf"
        invoice_obj.invoicePdf.save(pdf_name, ContentFile(pdf_buffer.getvalue()))
        invoice_obj.save()
        return f"Successfully generated PDF for Invoice {invoice_obj.invoiceNumber}"

    return f"Failed to generate PDF for Invoice {invoice_obj.invoiceNumber}"


def is_valid_contact(contact):
    return contact.isdigit() and 7 <= len(contact) <= 15


# Helper function to calculate SMS parts
def get_encoding_and_segments(text):
    gsm7_basic = "@£$¥èéùìòÇ\nØø\rÅåΔ_ΦΓΛΩΠΨΣΘΞ\x1bÆæßÉ !\"#¤%&'()*+,-./0123456789:;<=>?¡ABCDEFGHIJKLMNOPQRSTUVWXYZÄÖÑÜ§¿abcdefghijklmnopqrstuvwxyzäöñüà"
    gsm7_ext = "^{}\\[~]|€"
    encoding = "GSM-7"
    for char in text:
        if char not in gsm7_basic and char not in gsm7_ext:
            encoding = "UCS-2"
            break
    length = len(text)
    if encoding == "GSM-7":
        segments = 1 if length <= 160 else math.ceil(length / 153)
    else:
        segments = 1 if length <= 70 else math.ceil(length / 67)
    return encoding, segments, length


@shared_task
def process_campaign_contacts_task(
    campaign_id, file_path, contacts_string, user_id, message_text, client_id
):
    """
    Background worker to parse contacts, save them to the Campaign,
    and queue them up for the Vendor SMPP server.
    """
    message_text = re.sub(r"<[^>]+>", "", message_text)
    message_text = html.unescape(message_text).strip()

    createdContacts = []
    invalidContacts = []
    duplicateInInput = []
    seenInputs = set()

    try:
        campaign = Campaign.objects.get(id=campaign_id)
        user = User.objects.get(id=user_id)

        # ==========================================
        # 1. PARSE UPLOADED FILE (If exists)
        # ==========================================
        if file_path and default_storage.exists(file_path):
            file_name = file_path.lower()
            file_obj = default_storage.open(file_path)

            if file_name.endswith(".csv"):
                try:
                    decodedFile = file_obj.read().decode("utf-8")
                except UnicodeDecodeError:
                    file_obj.seek(0)
                    decodedFile = file_obj.read().decode("latin1")

                ioString = io.StringIO(decodedFile)
                reader = csv.DictReader(ioString)
                for row in reader:
                    row_lower = {k.lower(): v for k, v in row.items()}
                    contact = row_lower.get("contact", "").strip()
                    if contact and contact in seenInputs:
                        duplicateInInput.append(contact)
                    elif contact:
                        seenInputs.add(contact)
                        if is_valid_contact(contact):
                            createdContacts.append(
                                CampaignContact(
                                    campaign=campaign, contactNumber=contact
                                )
                            )
                        else:
                            invalidContacts.append(contact)

            elif file_name.endswith(".xlsx"):
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                headers = [
                    str(cell.value).lower()
                    for cell in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                contact_idx = headers.index("contact") if "contact" in headers else None

                if contact_idx is not None:
                    for row in ws.iter_rows(min_row=2):
                        contact = (
                            str(row[contact_idx].value).strip()
                            if row[contact_idx].value
                            else ""
                        )
                        if contact and contact in seenInputs:
                            duplicateInInput.append(contact)
                        elif contact:
                            seenInputs.add(contact)
                            if is_valid_contact(contact):
                                createdContacts.append(
                                    CampaignContact(
                                        campaign=campaign, contactNumber=contact
                                    )
                                )
                            else:
                                invalidContacts.append(contact)

        # ==========================================
        # 2. PARSE MANUAL TEXT INPUT (If exists)
        # ==========================================
        if contacts_string:
            contacts = [c.strip() for c in contacts_string.split(",") if c.strip()]

            for contact in contacts:
                if contact in seenInputs:

                    duplicateInInput.append(contact)

                else:
                    seenInputs.add(contact)

                    createdContacts.append(
                        CampaignContact(campaign=campaign, contactNumber=contact)
                    )

        # ==========================================
        # 3. SAVE TO DATABASE & QUEUE SMS
        # ==========================================
        if createdContacts:
            CampaignContact.objects.bulk_create(createdContacts)

            # 1. Setup Encoding & Fetch the Client
            encoding_type, total_segments, total_chars = get_encoding_and_segments(
                message_text
            )

            # ⚡️ Fetch the specific Client who is running this campaign
            client_obj = Client.objects.select_related("company").get(id=client_id)
            client_company = client_obj.company

            # 2. Loop to generate SMPP Messages
            for contact_obj in createdContacts:
                destination_addr = clean_phone_number(
                    contact_obj.contactNumber
                ).replace("+", "")
                print(
                    f"Queuing SMS to {destination_addr} with encoding {encoding_type} and {total_segments} segments"
                )
                unique_msg_id = str(uuid.uuid4())

                # --- A. FIND THE DESTINATION COUNTRY ---
                destination_country = None
                for i in range(4, 0, -1):
                    possible_code = destination_addr[:i]
                    destination_country = Country.objects.filter(
                        countryCode=possible_code, isDeleted=False
                    ).first()
                    if destination_country:
                        break

                if not destination_country:
                    print(f"Skipping {destination_addr}: Unrecognized Country Code")
                    continue

                # --- B. CALL THE ROUTING ENGINE ---
                # This uses the exact same logic as your SMPP Server!
                route_data, routing_error = get_route_and_cost(
                    client_obj, destination_country
                )

                if routing_error:
                    print(f"Routing Failed for {destination_addr}: {routing_error}")
                    # Optional: Create a FAILED SMSMessage here to track the failure
                    SMSMessage.objects.create(
                        client=client_obj,
                        systemId=client_obj.smppUsername,
                        destination=destination_addr,
                        text=message_text,
                        status="failed",
                        failure_reason=routing_error,
                        queued_at=timezone.now(),
                        failed_at=timezone.now(),
                    )
                    continue

                # --- C. CALCULATE COSTS ---
                total_vendor_cost = route_data["vendor_cost"] * total_segments
                total_client_cost = route_data["client_cost"] * total_segments

                vendor_obj = route_data["vendor"]
                smpp_obj = route_data["smpp"]
                raw_terminating_company = route_data["terminatingCompany"]

                print(f"Queuing SMS to {destination_addr} via {vendor_obj.profileName}")

                # --- D. THE ATOMIC LOCK (BILLING & SAVING) ---
                try:
                    with transaction.atomic():
                        # 1. LOCK ALL 3 BALANCES
                        locked_vendor_company = (
                            type(raw_terminating_company)
                            .objects.select_for_update()
                            .get(id=raw_terminating_company.id)
                        )
                        locked_client = Client.objects.select_for_update().get(
                            id=client_obj.id
                        )
                        locked_client_company = (
                            type(client_company)
                            .objects.select_for_update()
                            .get(id=client_company.id)
                        )

                        # 2. CHECK LIMITS (Pre-flight safety check)
                        if (
                            locked_vendor_company.usedVendorCredit + total_vendor_cost
                        ) > locked_vendor_company.vendorCreditLimit:
                            raise Exception("Insufficient Vendor Credit")
                        if (
                            locked_client.usedCredit + total_client_cost
                        ) > locked_client.creditLimit:
                            raise Exception("Insufficient Client Credit")
                        if (
                            locked_client_company.usedCustomerCredit + total_client_cost
                        ) > locked_client_company.customerCreditLimit:
                            raise Exception("Insufficient Company Credit")

                        # 3. DEDUCT ALL BALANCES
                        locked_vendor_company.usedVendorCredit += total_vendor_cost
                        locked_vendor_company.save(update_fields=["usedVendorCredit"])

                        locked_client.usedCredit += total_client_cost
                        locked_client.save(update_fields=["usedCredit"])

                        locked_client_company.usedCustomerCredit += total_client_cost
                        locked_client_company.save(update_fields=["usedCustomerCredit"])

                        # 4. CREATE THE SMS MESSAGE
                        parent_msg = SMSMessage.objects.create(
                            destination=destination_addr,
                            message_id=unique_msg_id,
                            text=message_text,
                            encoding=encoding_type,
                            segmentNumber=total_segments,
                            characterCount=total_chars,
                            status="queued",
                            vendor=vendor_obj,  # ⚡️ Assigned dynamically!
                            smpp=smpp_obj,  # ⚡️ Assigned dynamically!
                            client=locked_client,  # ⚡️ Client is now attached!
                            systemId=locked_client.smppUsername,
                            createdBy=user,
                            sendClientDlr=False,  # Campaigns don't need SMPP DLRs
                            queued_at=timezone.now(),
                        )
                        create_message_parts(parent_msg, message_text)

                        # 5. WRITE VENDOR LEDGER
                        VendorTransaction.objects.create(
                            vendor=vendor_obj,
                            message=parent_msg,
                            transactionType=TransactionType.DEDUCTION,
                            segments=total_segments,
                            ratePerSegment=route_data["vendor_cost"],
                            amount=total_vendor_cost,
                            balanceSpent=locked_vendor_company.usedVendorCredit,
                            description=f"Campaign - Routing charge for SMS {unique_msg_id}",
                        )

                        # 6. WRITE CLIENT LEDGER
                        ClientTransaction.objects.create(
                            client=locked_client,
                            message=parent_msg,
                            transactionType=TransactionType.DEDUCTION,
                            segments=total_segments,
                            ratePerSegment=route_data["client_cost"],
                            amount=total_client_cost,
                            chargePolicy=locked_client.invoicePolicy,
                            currency=(
                                locked_client.company.currency.name
                                if locked_client.company.currency
                                else "USD"
                            ),
                            balanceSpent=locked_client.usedCredit,
                            description=f"Campaign - Sent SMS {unique_msg_id}",
                        )

                        # 7. WRITE DETAILED REPORT
                        DetailedSMSReport.objects.create(
                            message=parent_msg,
                            text_message_id=unique_msg_id,
                            senderId=locked_client.smppUsername,
                            text=message_text,
                            part_total=total_segments,
                            client=locked_client.smppUsername,
                            clientRate=route_data["client_cost"],
                            client_charge=total_client_cost,
                            vendor=vendor_obj.profileName,
                            vendorRate=route_data["vendor_cost"],
                            vendor_charge=total_vendor_cost,
                            submitStatus="QUEUED",
                            operatorMNC=route_data.get("mnc", "Unknown"),
                            countryMCC=route_data.get("country_code", "Unknown"),
                            request_time=parent_msg.createdAt,
                            destination=destination_addr,
                        )
                except Exception as e:
                    print(f"Skipping {destination_addr} due to billing constraint: {e}")
                    # Save as failed due to credit
                    SMSMessage.objects.create(
                        client=client_obj,
                        systemId=client_obj.smppUsername,
                        destination=destination_addr,
                        text=message_text,
                        status="failed",
                        failure_reason=str(e),
                        queued_at=timezone.now(),
                        failed_at=timezone.now(),
                    )

    except Exception as e:
        print(f"🔥 Task Failed for Campaign {campaign_id}: {str(e)}")

    finally:
        # CLEANUP: Delete the temp file so your server doesn't bloat
        if file_path and default_storage.exists(file_path):
            try:
                if "file_obj" in locals():
                    file_obj.close()
            except Exception:
                pass
            default_storage.delete(file_path)
